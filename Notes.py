import openpyxl as oxl
import re
import openpyxl
import numpy as np
import pymongo
import pprint
import datetime

# wb = oxl.load_workbook('DK__AtlasCopCo_Miniglobal_WP237689.xlsx')
# wb = oxl.load_workbook('IT_Etro_Spa__Renewal_WP239809.xlsx')
# wb = oxl.load_workbook('ES_WP240238__RULES__AB_BIOTICS_WP240238.xlsx')
# wb = oxl.load_workbook("GB_Ineos_NYZ3_LocCorp_WP237164.xlsx")

def notes2(excel_path, corporate_name, db, system_date, system_time, file_date, file_time, fare_sheet_names=None):
    wb_obj = excel_path
    wb = oxl.load_workbook(excel_path)
    pos_sheet_notes = []
    pos_sheet_rules = []

    def possheetnotes(list11):
        noteslist=[]
        noteslistf=[]
        for i in range(0,len(list1)):
            if list1[i] =="Notes":
                for j in range(i,len(list1)):            
                    noteslist.append(list1[j])
        #     print(noteslist)
        for i in range(0,len(noteslist)):
            if noteslist[i] =="Markets Rule Waiver " or noteslist[i]=="Markets Rule Waiver  - ITC5 ":
                break
            else:
                noteslistf.append(noteslist[i])
            noteslistf1 = noteslistf[1:]
        pos_notes = [{"field_key": "notes", "field_value": k} for k in noteslistf1]
        return pos_notes

    for sheet in wb:    
        # if sheet.title.startswith("POO BRU With Market Rules"):
            # sheet_t = sheet.title
        if sheet.title in fare_sheet_names:
            list1=[]
            unhiddenrow,unhiddencol=[],[]
            for rowNum,rowDimension in sheet.row_dimensions.items():
                if rowDimension.hidden == True:
                    unhiddenrow.append(rowNum)
            for row in range(1,sheet.max_row+1):
                if row not in unhiddenrow:
                    for col in range(1,sheet.max_column+1):
                        cell=sheet.cell(row=row,column=col)
                        if cell.value!=None:
                            list1.append(cell.value)
            list1 = [str(x) for x in list1]
            for val in list1:
                if val.startswith("Market Rules") or val.startswith("Markets Rule Waiver ") or val.startswith("BE1 Corp Fare Rules ( Exceptions to Base Rule )")or val.startswith("Markets Rule Waiver  - ITC5 "):
                    list2 = list1
                    listMR = []

                    for i in range(0,len(list1)):
                        if list1[i] =="Markets Rule Waiver "  or list1[i]=="Markets Rule Waiver  - ITC5 "  or list1[i]==  "BE1 Corp Fare Rules ( Exceptions to Base Rule )":
                            for j in range(i,len(list1)):            
                                listMR.append(list1[j])   
                            listMR1 = listMR[2:]
                    pos_sheet_rules = [{"field_key": listMR[1], "field_value": k} for k in listMR1]
                    print(pos_sheet_rules)
                    if "Notes" in list1:

                        pos_sheet_notes = possheetnotes(list1)
    #                     print("##@#@##############/n######################/n#######")  
    #                     print("##@#@##############/n######################/n#######")       
                        print(pos_sheet_notes)
                    else:
                        pos_sheet_notes = []
    #                     print("##@#@##############/n######################/n#######")  
    #                     print("##@#@##############/n######################/n#######")
        
             
                    if "General Notes" in list1:

                        pos_sheet_notes = []
                    else:
                        general_notes = []        
                    print(general_notes)  

                    if "General " in list1:

                        pos_sheet_notes = []
                    else:
                        general = []                  

                else:
                    continue

    _id = db.corporate.rulesheet.update(
        {
            "corporate_name": corporate_name,
            "file_date": file_date,
            "file_time": file_time
        },
        {
            "$set": {
                "pos_sheet_notes": pos_sheet_notes,
                "pos_sheet_rules": pos_sheet_rules
            }
        },
        upsert=True
        )

    if len(pos_sheet_rules) > 0:
        return True, _id
    else:
        return False, _id

def rules(excel_path, corporate_name, db, system_date, system_time, file_date, file_time, pos_sheet_rules_exists):
    xlsx_file = excel_path
    wb_obj = openpyxl.load_workbook(xlsx_file)
    markets_with_rule_waiver_list = []
    general_notes = []
    general = []

    for sheet in wb_obj:
        if "Notes" in sheet.title:
            col_names = []

            list1= []
            notes = []
            notes2= []
            poolist=[]
            marketlist = []
            # j = 0
            cat16_break = False
            for row in range(1,sheet.max_row+1):
                for col in range(1,sheet.max_column+1):
                    cell=sheet.cell(row=row,column=col)
                    if cell.value!=None:
                        list1.append(cell.value)


            dict1={}
            general_notes = []
            dict1["general"] = {}
            dict1['General_Notes'] = {}
            dict1['General_Notes']['Fares_&_Ticketing_Particulars'] = {}
            dict1['Markets_with_Rule_Waiver'] = {}
            lcfn_patterns = {

            'corporate_revenue_p': ['Corporate Revenue'],
            'corporate_target_p':  ['Corporate Target'],
            'corporate_share_p':  ['Corporate Share'],
            'pax_eligibility_p': ['Pax Eligibility'],
            'rfp_pricing_annex_p': ['RFP Pricing Annex']

            }
            general_notes_patterns = {
                
            'resultant_fares_p': ['Resultant Fares'],
            'rule_conditions_p': ['Rule Conditions ','Rule Conditions'],
            'child/infant_discounts_p': ['Child / Infant Discounts'],
            'general_exclusions_p': ['General Exclusions'],
            'ticket_fare_box_p': ['Ticket Fare Box'],
            'non-discounted_fares_p': ['Non-Discounted Fares'],
            'Ticket_Designator/Tour_Code_p': ['Ticket Designator/Tour Code'],
            'market_commission_p': ['Market Commission'],
            'market_fuel_surcharge_p': ['Market Fuel Surcharge']


            }
            mwrw_patterns = {
                
            'cat_06_p':['Cat 06.', '06. MINIMUM STAY'],
            'cat_16_p':['Cat 16.', '16. PENALTIES'],
            'cat_07_p': ['Cat 07.', '07. MAXIMUM STAY'],
            'cat_05_p': ['Cat 05', '05. ADVANCE']

            }
            markets_with_rule_waiver_list = []
            for i  in range(len(list1)):

                if "Local Corporate Framework Notes" in str(list1[i]):
                    

                    
                    general = []
                    dict1['Local_Corporate_Framework_Notes'] = {}
                    
                    for i  in range(len(list1)):

                        if list1[i] in lcfn_patterns['corporate_revenue_p']:
                            dict1["Local_Corporate_Framework_Notes"]['corporate_revenue']=list1[i+1]
                            general.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})
                            
                        if list1[i] in lcfn_patterns['corporate_target_p']:
                            dict1["Local_Corporate_Framework_Notes"]['corporate_target']=list1[i+1]
                            general.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})

                        if list1[i] in lcfn_patterns['corporate_share_p']:
                            dict1["Local_Corporate_Framework_Notes"]['corporate_share']=list1[i+1]
                            general.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})

                        if list1[i] in lcfn_patterns['pax_eligibility_p']:
                            dict1["Local_Corporate_Framework_Notes"]['pax_eligibility']=list1[i+1]
                            general.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})

                        if list1[i] in lcfn_patterns['rfp_pricing_annex_p']:
                            dict1["Local_Corporate_Framework_Notes"]['rfp_pricing_annex']=list1[i+1]
                            general.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})

                    print(general)
                    # db.corporate.rulesheet.insert_one({'local_corporate_framework_notes':dict1['Local_Corporate_Framework_Notes']})
                        

            for i  in range(len(list1)):

                if list1[i] in general_notes_patterns['resultant_fares_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['resultant_fares']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})


                if list1[i] in general_notes_patterns['rule_conditions_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['rule_conditions']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().strip().replace(" ", "_"), "field_value":list1[i+1]})
                    
                if list1[i] in general_notes_patterns['child/infant_discounts_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['child/infant_discounts']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace("/", "_").replace(" ", "_").replace("__", ""), "field_value":list1[i+1]})
                    

                if list1[i] in general_notes_patterns['general_exclusions_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['general_exclusions']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})

                if list1[i] in general_notes_patterns['ticket_fare_box_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['ticket_fare_box']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})
                   


                if list1[i] in general_notes_patterns['non-discounted_fares_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['non-discounted_fares']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace(" ", "_").replace("-", "_"), "field_value":list1[i+1]})
                    

                if list1[i] in general_notes_patterns['Ticket_Designator/Tour_Code_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['Ticket_Designator/Tour_Code']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})
                    

                if list1[i] in general_notes_patterns['market_commission_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['market_commission']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})


                if list1[i] in general_notes_patterns['market_fuel_surcharge_p']:
                    dict1['General_Notes']["Fares_&_Ticketing_Particulars"]['market_fuel_surcharge']=list1[i+1]
                    general_notes.append({"field_key":list1[i].lower().replace(" ", "_"), "field_value":list1[i+1]})



                if "Markets with Rule Waiver" in str(list1[i]):

                    
                    marketlist = list1[i:]
                    dict1['Markets_with_Rule_Waiver'] = {}

                    for i  in range(len(marketlist)):
                        if "POO" in str(marketlist[i]):
                            poolist.append(i)
                    for i in range(len(poolist)):

                        j = len(poolist) - 1
                        
                        if i != j :
                            sas = marketlist[poolist[i]:poolist[i+1]]
                            j = marketlist[poolist[i]]
                            dict1['Markets_with_Rule_Waiver'][j] = {}
                            notes.append(j)
                            for i in range(len(sas)):
                                markets_with_rule_waiver_list.append({"field_key":j.lower().replace(" ", "_"), "field_value":sas[i]})
                            for i in range(len(sas)):
                                for patterns in mwrw_patterns:
                                    for pattern in mwrw_patterns[patterns]:
                                        if pattern in str(sas[i]):
                                            if patterns == 'cat_16_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_16'] = sas[i:]
                                                notes.extend(sas[i:])
                                            if patterns == 'cat_05_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_05']=sas[i], sas[i+1]
                                                notes.append(sas[i])
                                                notes.append(sas[i+1])

                                            if patterns == 'cat_06_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_06']=sas[i], sas[i+1]
                                                notes.append(sas[i])
                                                notes.append(sas[i+1])
                                
                                            if patterns == 'cat_07_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_07']=sas[i], sas[i+1]
                                                notes.append(sas[i])
                                                notes.append(sas[i+1])

                                    
                                    
                                notes2 = [item for item in sas if item not in notes]
                                

                                dict1['Markets_with_Rule_Waiver'][j]['notes']=notes2
                                    

                                

                                

                        elif i == j :

                            j = marketlist[poolist[i]]
                            dict1['Markets_with_Rule_Waiver'][j] = {}
                            sas = marketlist[poolist[i]:]
                            notes.append(j)
                            for i in range(len(sas)):
                                markets_with_rule_waiver_list.append({"field_key":j.lower().replace(" ", "_"), "field_value":sas[i]})
                            for i in range(len(sas)):
                                for patterns in mwrw_patterns:
                                    for pattern in mwrw_patterns[patterns]:
                                        if pattern in str(sas[i]):
                                            if patterns == 'cat_16_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_16'] = sas[i:]
                                                notes.extend(sas[i:])
                                            if patterns == 'cat_05_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_05']=sas[i], sas[i+1]
                                                notes.append(sas[i])
                                                notes.append(sas[i+1])

                                            if patterns == 'cat_06_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_06']=sas[i], sas[i+1]
                                                notes.append(sas[i])
                                                notes.append(sas[i+1])
                                
                                            if patterns == 'cat_07_p':
                                                dict1['Markets_with_Rule_Waiver'][j]['cat_07']=sas[i], sas[i+1]
                                                notes.append(sas[i])
                                                notes.append(sas[i+1])


                                notes2 = [item for item in sas if item not in notes]

                                dict1['Markets_with_Rule_Waiver'][j]['notes']=notes2

                                # db.corporate.rulesheet.insert_one({'markets_with_rule_waiver':dict1['Markets_with_Rule_Waiver']})
            pprint.pprint(markets_with_rule_waiver_list)

            print(general_notes)


    data_to_insert = dict()
    data_to_insert["general_notes"] = general_notes
    data_to_insert["general"] = general
    if not pos_sheet_rules_exists:
        data_to_insert["pos_sheet_rules"] = markets_with_rule_waiver_list

    # Creation of Dummy id
    import datetime
    import time
    dt = datetime.datetime.now()
    unix_time = (time.mktime(dt.timetuple()))
    data_to_insert["dummy_id"] = str(int(unix_time))
    db.corporate.rulesheet.update(
        {
            "corporate_name": corporate_name,
            "file_date": file_date,
            "file_time": file_time
        },
        {
            "$set": data_to_insert
            # "$set": {
            #         "general_notes": general_notes,
            #         "markets_with_rule_waiver": markets_with_rule_waiver_list,
            #         "general": general
            #     }
        },
        upsert=True
        )
    return str(int(unix_time))
    # db.corporate.rulesheet.insert_one(
    #     {
    #     "general_notes": general_notes,
    #     "markets_with_rule_waiver": markets_with_rule_waiver_list,
    #     "corporate_name":corporate_name,
    #     "system_date":s_date,
    #     "system_time":s_time,
    #     "file_date":f_date,
    #     "file_time":f_time
    #         # "general": dict1["general"],
    #         # "general_notes": dict1["General_Notes"]["Fares_&_Ticketing_Particulars"],
    #         # "markets_with_rule_waiver": dict1["Markets_with_Rule_Waiver"]
    #     })

def parse_notes(excel_path, corporate_name, db, SYSTEM_DATE, SYSTEM_TIME, FILE_DATE, FILE_TIME):

    s_date = SYSTEM_DATE
    s_time = SYSTEM_TIME
    f_date = FILE_DATE
    f_time = FILE_TIME
    # def possheetnotes(list11, excel_path):
    pos_sheet_rules_exists = notes2(excel_path)
    rules(excel_path, corporate_name, db, s_date, s_time, f_time, f_date, pos_sheet_rules_exists)

if __name__=='__main__':
    client = pymongo.MongoClient(
        '3.6.201.161:27022',
        username='data.EK',
        password='data@123',
        authSource='admin',
        authMechanism='SCRAM-SHA-1'
    )
    db = client['rawDB_prod']

    excel_path = 'Ericsson_Global Deal.xlsx'
    corporate_name = "Occidental"
    system_date = datetime.date.today().strftime("%Y-%m-%d")
    now = datetime.datetime.now()
    system_time = now.strftime("%H:%M:%S")
    file_time = "13:50:45"
    file_date = "2020-08-11"

    parse_notes(excel_path, corporate_name, db, system_date, system_time, file_date, file_time)
    

