from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import pprint
import re
import os
import datetime
from pymongo import MongoClient
from bson.objectid import ObjectId
from collections import Counter

import pandas as pd 

# SYSTEM_DATE = datetime.date.today().strftime("%Y-%m-%d")

def get_db_client():

    client = MongoClient(
            '3.6.201.161:27022',
            username='data.EK',
            password='data@123',
            authSource='admin',
            #authMechanism='SCRAM-SHA-256'
        )

    return client

def get_coords_def_type1(ws, raw_db):

    corp_framework_coords = dict()
    corp_framework_coords['fare_class'] = None
    fare_class_coords = dict()
    components_coords = dict()

    fare_sheet_keyword_patterns = raw_db.corporate.patterns.find_one(
            {"flag": "keyword_pattern"}, 
            {
                "fare_sheet.type1_fare_class_regex_pattern": 1, 
                "fare_sheet.type1_fare_brand_regex_pattern": 1,
                "fare_sheet.type1_origin_dest_regex_patterns": 1,
                "fare_sheet.type1_components_regex_patterns": 1,
                "fare_sheet.type1_percent_discount_published_ending_regex_patterns": 1,
                "_id": 0})["fare_sheet"]
                
    # fare_class_regex_patterns = {
    # 'first_class' : '\s*first\s*class',
    # 'business_class' : '\s*business\s*class',
    # 'economy_class' : '\s*economy\s*class',
    # 'all_cabin' :  '\s*all\s*cabin\^?'
    # }
    # print(fare_class_regex_patterns)
    fare_class_regex_patterns = fare_sheet_keyword_patterns["type1_fare_class_regex_pattern"]
    # print(fare_class_regex_patterns)

    # fare_brand_regex_patterns = {
    #     'flex_plus': '\s*flex\s*plus',
    #     'flex_cap': '\s*flex\^',
    #     'saver': '\s*saver',
    #     'flex': '\s*flex',
    #     'special': 'special'
    # }
    # print(fare_brand_regex_patterns)
    fare_brand_regex_patterns = fare_sheet_keyword_patterns["type1_fare_brand_regex_pattern"]
    # print(fare_brand_regex_patterns)

    # origin_dest_regex_patterns = {
    #     'origin': ['\s*origin', '\s*poo'],
    #     'destination': ['\s*destination'],
    # }
    # print(origin_dest_regex_patterns)
    origin_dest_regex_patterns = fare_sheet_keyword_patterns["type1_origin_dest_regex_patterns"]
    # print(origin_dest_regex_patterns)

    # components_regex_patterns = {
    #     'ow_rt': ['\s*fare\s*type','\s*rt/ow'],
    #     'fare_box': ['\s*fare\s*box'],
    #     'tkt_fbc_code': ['\s*tktd\s*fbasis\s*code'],
    #     # 'discounts_applicable_on_published_first_business_economy_class_fare_basis': ['\s*discounts\s*applicable\s*on\s*published\s*first(.*?)business(.*?)(and)?\s*economy'],
    #     'disc_applicable': ['\s*discounts\s*applicable\s*on\s*published\s*first(.*?)business(.*?)(and)?\s*economy'],
    #     'tier': ['\s*tier\s*', '\s*corporate\s*deal\s*tier'],
    #     # 'revenue_qualifiers_range': ['\s*revenue\s*qualifiers\s*range'],
    #     'revenue_range': ['\s*revenue\s*qualifiers\s*range'],
    #     'pax_type': ['pax\s*type'],
    #     'tkt_designator': ['ticket designator']
    # }
    # print(components_regex_patterns)
    components_regex_patterns = fare_sheet_keyword_patterns["type1_components_regex_patterns"]
    # print(components_regex_patterns)

    # percent_discount_published_ending_regex_patterns = ['percentage\s*discount\s*off\s*published\s*(fares)?\s*ending']
    # print(percent_discount_published_ending_regex_patterns)
    percent_discount_published_ending_regex_patterns = fare_sheet_keyword_patterns["type1_percent_discount_published_ending_regex_patterns"]
    # print(percent_discount_published_ending_regex_patterns)

    fare_class_coords_set = False
    fare_brand_coords_set = False
    origin_dest_coords_set = False
    rbd_coords_set = False


    for row_cells in ws.iter_rows(min_row = 1, max_row=15):
        is_fare_class_row = False
        is_fare_brand_row = False
        prev_class = None
        
        if fare_brand_coords_set and fare_class_coords_set and origin_dest_coords_set:
            break
        
        for cell in row_cells:
            value = cell.value
            is_assigned = False
            
            if not origin_dest_coords_set:
                if isinstance(value, str):
                    is_matched = False
                    for patterns in origin_dest_regex_patterns:
                        for pattern in origin_dest_regex_patterns[patterns]:
                            if (re.match(pattern, value, flags=re.IGNORECASE) is not None):
                                is_matched = True
                                is_assigned = True
                                if patterns == 'origin':
                                    corp_framework_coords['origin'] = dict()
                                    corp_framework_coords['origin']['col'] = get_column_letter(cell.column)

                                elif patterns == 'destination':
                                    corp_framework_coords['destination'] = dict()
                                    corp_framework_coords['destination']['col'] = get_column_letter(cell.column)
                                    origin_dest_coords_set = True
                                    
                            if is_matched:
                                break

            if isinstance(value, str):
                is_matched = False
                for pattern in percent_discount_published_ending_regex_patterns:
                    if (re.match(pattern, value, flags=re.IGNORECASE) is not None):
                        is_matched = True
                        is_assigned = True
                        corp_framework_coords['sheet_desc'] = value

                
            # parsing fare_class coordinates
            if not fare_class_coords_set:
                if isinstance(value, str):
                    is_matched = False
                    for fc in fare_class_regex_patterns:
                        if (re.match(fare_class_regex_patterns[fc], value, flags=re.IGNORECASE) is not None):
                            is_fare_class_row = True
                            is_matched = True
                            is_assigned = True
                            cur_col = cell.column
                            
                            
                            if fc not in fare_class_coords:
                                fare_class_coords[fc] = dict()
                                fare_class_coords[fc]['start_col'] = cur_col
                            if prev_class != None:
                                fare_class_coords[prev_class]['end_col'] = cur_col
                            prev_class = fc
                        if is_matched:
                            break
            
            # parsing fare_brand coordinates
            elif not fare_brand_coords_set:
                if isinstance(value, str):
                    is_matched = False
                    for fb in fare_brand_regex_patterns:
                        if (re.match(fare_brand_regex_patterns[fb], value, flags=re.IGNORECASE) is not None):
                            is_fare_brand_row = True
                            is_matched = True
                            is_assigned = True
                            cur_col = cell.column
                            fare_cls = None
                            
                            fb1 = None
                            rbd1 = None

                            if "/" in value:
                                fb_split = value.split("/")
                                if (re.match(fare_brand_regex_patterns[fb], fb_split[0], flags=re.IGNORECASE) is not None):
                                    fb1 = fb_split[1].strip().lower().replace(" ", "_")
                                else:
                                    fb1 = fb_split[0].strip().lower().replace(" ", "_")


                            for fare_class in fare_class_coords:
                                valid = False
                                start_col = None
                                end_col = None
                                
                                if 'start_col' in fare_class_coords[fare_class]:
                                    start_col = fare_class_coords[fare_class]['start_col']
                                if 'end_col' in fare_class_coords[fare_class]:
                                    end_col = fare_class_coords[fare_class]['end_col']
                                
                                if start_col != None and end_col != None and cur_col >= start_col and cur_col < end_col:
                                    valid = True
                                elif end_col == None and cur_col >= start_col:
                                    valid = True
                                
                                if valid:
                                    fare_cls = fare_class
                                    # print(fb + ' in col ' + str(cur_col) + ' belong to ' + fare_cls)
                                    rbd_coord = (cell.row + 1, cell.column)

                                    rbd_value = ws.cell(row = rbd_coord[0], column = rbd_coord[1]).value.strip().upper()
                                    rbd_value = rbd_value.replace(",", "/")

                                    if fb not in fare_class_coords[fare_cls]:
                                        fare_class_coords[fare_cls][fb] = dict()

                                    if fb1 is not None and fb1 not in fare_class_coords[fare_cls]:
                                        fare_class_coords[fare_cls][fb1] = dict()

                                    if fb1 is not None:
                                        rbd_split = rbd_value.split("\n")
                                        for rbd_detail in rbd_split:
                                            c = rbd_detail.split(":")
                                            if c[0].strip().lower().replace(" ", "_") == fb:
                                                rbd_value = c[1].strip()
                                            elif c[0].strip().lower().replace(" ", "_") == fb1:
                                                rbd1 = c[1].strip()

                                    counter = 1
                                    while(rbd_value in fare_class_coords[fare_cls][fb]):
                                        rbd_value = rbd_value + str(counter)
                                        counter += 1
                                    fare_class_coords[fare_cls][fb][rbd_value] = {'col': get_column_letter(cur_col)}

                                    if fb1 is not None:
                                        fare_class_coords[fare_cls][fb1][rbd1] = {'col': get_column_letter(cur_col)}
                                    if 'start_row' not in fare_class_coords:
                                        corp_framework_coords['start_row'] = cell.row + 2

                                    fb_col_counter = 1
                                    while(ws.cell(row=cell.row, column=cell.column+fb_col_counter).value == None and end_col != None and cell.column+counter <= end_col):
                                        rbd_coord = (cell.row + 1, cell.column+fb_col_counter)
                                        rbd_value = ws.cell(row = rbd_coord[0], column = rbd_coord[1]).value.strip().upper()
                                        rbd_value = rbd_value.replace(",", "/")
                                        while(rbd_value in fare_class_coords[fare_cls][fb]):
                                            rbd_value = rbd_value + str(counter)
                                            counter += 1
                                        fare_class_coords[fare_cls][fb][rbd_value] = {'col': get_column_letter(rbd_coord[1])}
                                        fb_col_counter += 1

                                    break
                        if is_matched:
                            break
            
            if isinstance(value, str) and not is_assigned:
                is_matched = False
                for component in components_regex_patterns:
                    for comp_pat in components_regex_patterns[component]:
                        if (re.match(comp_pat, value, flags=re.IGNORECASE) is not None):
                            is_matched = True
                            components_coords[component] = {'col': get_column_letter(cell.column)}
                            break
                    if is_matched:
                        break
                
        if is_fare_class_row:
            fare_class_coords_set = True
        if is_fare_brand_row:
            fare_brand_coords_set = True
                           
    corp_framework_coords['fare_class'] = fare_class_coords
    corp_framework_coords['components'] = components_coords

    return corp_framework_coords

def parse_corp_framework_sheet_type1(sheet, coords_def):

    complete_sheet_data = list()
    data_columns = dict()
    comp_columns = dict()
    comp_details = list()
    discount_details = list()
    isFareTypeGivenHorizontally = False
    fare_type_row = None
    table_data = dict()
    extra_cols = list()
    ow_rt_vals = dict()

    for cls in coords_def['fare_class']:
        for fb in coords_def['fare_class'][cls]:
            if fb in ['start_col', 'end_col']:
                continue
            for rbd in coords_def['fare_class'][cls][fb]:
                key = cls + '||' + fb + '||' + rbd
                col = coords_def['fare_class'][cls][fb][rbd]['col']
                data_columns[key] = col

                fare_class_ = cls.lower().strip()
                fare_brand_ = fb.lower().strip()
                fare_family = ""
                if "*" in rbd:
                    fare_family = "*" + rbd.split("*")[1]
                    rbd = rbd.replace(fare_family, "")
                separator = "/"
                if "," in rbd:
                    separator = ","
                rbd_list = [r.lower().strip() + fare_family for r in rbd.split(separator)]

                if fare_class_ not in table_data:
                    table_data[fare_class_] = dict()
                if fare_brand_ not in table_data[fare_class_]:
                    table_data[fare_class_][fare_brand_] = list()
                table_data[fare_class_][fare_brand_].extend(rbd_list)


    for comp in coords_def['components']:
        comp_columns[comp] = coords_def['components'][comp]['col']
        extra_cols.append(comp)
                
                
    origin_col = coords_def['origin']['col']
    dest_col = coords_def['destination']['col']

    print('column details of respective rbds, fbs and fare_classes')
    print('******' * 10)
    pprint.pprint(data_columns)
    print('******' * 10)
    print('components details')
    pprint.pprint(comp_columns)
    print('******' * 10)
     
    start_row = coords_def['start_row']
    cur_row = start_row
    prev_origin = None
    prev_dest = None
    prev_comp = dict()

    while(True):
        break_condition = 0
        value_dict = dict()

        dest_val = sheet[dest_col + str(cur_row)].value
        if (dest_val == None):
            if sheet[dest_col + str(cur_row + 1)].value == None:
                break_condition += 1
            else:
                cur_row = cur_row + 1
                continue
                
        origin_val = sheet[origin_col + str(cur_row)].value
        if origin_val == None:
            break_condition += 1
        if break_condition == 2:
            break
        
        if origin_val == '(City code  or Country name)':
            isFareTypeGivenHorizontally = True
            fare_type_row = str(cur_row)
            cur_row += 1
            continue
        
        value_dict['origin'] = origin_val if (origin_val != None) else prev_origin
        value_dict['destination'] = dest_val if (dest_val != None) else prev_dest


        value_dict['discount_details'] = list()
        if origin_val != None:
            prev_origin = origin_val
        if dest_val != None:
            prev_dest = dest_val 

        prev_discount = None
        for column in data_columns:
            col = data_columns[column]
            (cls, fb, rbd) = column.split('||')
            cur_discount_dict = dict()
            discount_raw = sheet[col + str(cur_row)].value 
            print(discount_raw, type(discount_raw))
            if discount_raw != None and isinstance(discount_raw, str) and "_x000D_" in discount_raw:
                discount_raw = discount_raw.replace("_x000D_", "")
            if isinstance(discount_raw, float):
                discount_raw = round(discount_raw * 100, 1)

            cur_discount_dict = {
                'fare_class': cls,
                'fare_brand': fb,
                'rbd': rbd,
                'discount': discount_raw if discount_raw != None else prev_discount
            }
            if isFareTypeGivenHorizontally:
                cur_discount_dict['fare_type'] = sheet[col + fare_type_row].value
                if "*" in rbd:
                    fare_family = "*" + rbd.split("*")[1]
                    rbd = rbd.replace(fare_family, "")
                rbd_list = [r.lower().strip() + fare_family for r in rbd.split('/')]
                for rbd_ in rbd_list:
                    ow_rt_vals[cls + "|" + fb + "|" + rbd_] = sheet[col + fare_type_row].value

                
            value_dict['discount_details'].append(cur_discount_dict)
            
            # if sheet[col + str(cur_row)].value != None:
            if discount_raw != None:
                prev_discount = discount_raw

        for column in comp_columns:
            comp_col = comp_columns[column]
            
            if column not in prev_comp:
                prev_comp[column] = sheet[comp_col + str(cur_row)].value
                
            value_dict[column] = sheet[comp_col + str(cur_row)].value if sheet[comp_col + str(cur_row)].value != None else prev_comp[column]
            
            if sheet[comp_col + str(cur_row)].value != None:
                prev_comp[column] = sheet[comp_col + str(cur_row)].value

        value_dict["remarks"] = list()
        if "sheet_desc" in coords_def:
            value_dict["remarks"].append(coords_def["sheet_desc"])

        discount_details.append(value_dict)
        print(value_dict)
        print('*****' * 10)
        cur_row = cur_row + 1
        complete_sheet_data.append(value_dict)

    # print('table data')
    # pprint.pprint(table_data)
    # print("extra columns: ")
    # print(extra_cols)
    # print("ow_rt_row-wize")
    # print(ow_rt_vals)

    table_data_ = list()
    for cls in table_data:
        cur_class_dict = dict()
        cur_class_dict['header'] = cls 
        cur_class_dict['subheaders'] = list()
        cur_class_dict['span'] = list()
        cur_class_dict['rbdlist'] = list()

        for fb in table_data[cls]:
            cur_class_dict['subheaders'].append(fb)
            cur_class_dict['span'].append(len(table_data[cls][fb]))
            rbd_list = table_data[cls][fb]
            if isFareTypeGivenHorizontally:
                for i, rbd_ in enumerate(rbd_list):
                    rbd_list[i] = rbd_ + "\n" + ow_rt_vals[cls + "|" + fb + "|" + rbd_]
            cur_class_dict['rbdlist'].extend(rbd_list)
        table_data_.append(cur_class_dict)

    # print('table data trans')
    # pprint.pprint(table_data_)

    header_details = {
        "extra_columns": extra_cols,
        "table_data": table_data_
    }
    print("header details: ")
    pprint.pprint(header_details)

    return complete_sheet_data, header_details

def get_coords_def_type2(ws, raw_db):

    print("parsing type2 sheet coordinates")
    print("****" * 10)
    
    fare_sheet_keyword_patterns = raw_db.corporate.patterns.find_one(
        {"flag": "keyword_pattern"},
        {
            "_id": 0,
            "fare_sheet.type2_discount_applied_fare_family_regex_patterns": 1,
            "fare_sheet.type2_discount_percentage_regex_patterns": 1,
            "fare_sheet.type2_origin_dest_regex_patterns": 1,
            "fare_sheet.type2_components_regex_patterns": 1,   
        }
    )["fare_sheet"]
    # discount_applied_fare_family_regex_patterns = [
    # 'fare\s*family\s*to\s*which\s*the\s*discount\s*to\s*be\s*applied',
    # 'discounts\s*applicable\s*on\s*published\s*business\s*and\s*economy\s*class\s*fare\s*basis'
    # ]
    # print(discount_applied_fare_family_regex_patterns)
    discount_applied_fare_family_regex_patterns = fare_sheet_keyword_patterns["type2_discount_applied_fare_family_regex_patterns"]
    # print(discount_applied_fare_family_regex_patterns)

    # discount_percentage_regex_patterns = ['discount\s*percentage\s*(/\s*)?fixed\s*amount', 'percentage\s*discount']
    # print(discount_percentage_regex_patterns)
    discount_percentage_regex_patterns = fare_sheet_keyword_patterns["type2_discount_percentage_regex_patterns"]
    # print(discount_percentage_regex_patterns)

    # origin_dest_type2_regex_patterns = {
    #     'origin': ['travel\s*origin(ation)?'],
    #     'destination': ['travel\s*destination(nation)?'],
    # }
    # print(origin_dest_type2_regex_patterns)
    origin_dest_type2_regex_patterns = fare_sheet_keyword_patterns["type2_origin_dest_regex_patterns"]
    # print(origin_dest_type2_regex_patterns)

    # components_type2_regex_patterns = {
    #     'ow/rt': ['\s*fare\s*type','rt(\s*)/(\s*)ow', 'ow(\s*)/(\s*)rt'],
    #     'fare_box': ['\s*fare\s*box'],
    #     'tktd_fb_code': ['\s*tktd\s*fbasis\s*code', '\s*ticketed\s*fare\s*basis\s*code'],
    #     'base_rule': ['base\s*(fare)?\s*rule'],
    #     'tour_code': ['tour\s*code'],
    #     'remarks': ['remarks'],
    #     'ticket_designator': ['ticket\s*designator']
    # }
    # print(components_type2_regex_patterns)
    components_type2_regex_patterns = fare_sheet_keyword_patterns["type2_components_regex_patterns"]
    # print(components_type2_regex_patterns)

    corp_framework_coords_type2 = dict()
    fare_family_cols = dict()
    discount_percent_cols = dict()
    components_coords_type2 = dict()
    baseheaders = list()
    component_list = list()
    base_spans = list()
    sub_spans = list()
    key_values = list()
    fare_family_list = list()
    fare_brand_list = list()
    seasonality_list = list()

    origin_destination_coords_type2_set = False
    fares_family_coords_set = False
    fares_discount_coords_set = False

    for row_cells in ws.iter_rows(min_row = 1, max_row=80):
        for cell in row_cells:
            
            value = cell.value
            if isinstance(value, str):
                print(value)
                print("****")
                for key in origin_dest_type2_regex_patterns:
                    for pattern in origin_dest_type2_regex_patterns[key]:
                        if (re.search(pattern, value, flags=re.IGNORECASE) is not None):
                            baseheaders.append(key)
                            corp_framework_coords_type2[key] = {'col': get_column_letter(cell.column)}
                            origin_destination_coords_type2_set = True
                            if 'start_row' not in corp_framework_coords_type2:
                                end_row = cell.row + 1
                                while(ws.cell(row=end_row, column=cell.column).value == None):
                                    end_row += 1
                                corp_framework_coords_type2['start_row'] = end_row
                            
                if origin_destination_coords_type2_set:
                    for pattern in discount_applied_fare_family_regex_patterns:
                        if (re.search(pattern, value, flags=re.IGNORECASE) is not None):
                            baseheaders.append("Fare Family to which the discount to be applied")
                            start_col = cell.column
                            end_col = start_col + 1
                            print("in fare family block")
                            print(start_col, end_col)
                            while(ws.cell(row=cell.row, column = end_col).value == None):
                                end_col = end_col + 1
                            print(start_col, end_col)
                            for i in range(start_col, end_col):
                                fare_class = ws.cell(row=cell.row + 1, column = i).value.strip()
                                fare_class = re.sub('\s+(-\s+)?', ' ', fare_class)
                                fare_family_list.append(fare_class)
                                fare_family_cols[fare_class] = {'col': get_column_letter(i)}
                            if 'start_row' not in corp_framework_coords_type2:
                                    corp_framework_coords_type2['start_row'] = cell.row + 2
                            fares_family_coords_set = True
                            
                if fares_family_coords_set:
                    for pattern in discount_percentage_regex_patterns:
                        if (re.search(pattern, value, flags=re.IGNORECASE) is not None):
                            baseheaders.append("Percentage Discount")
                            start_col = cell.column
                            end_col = start_col + 1
                            while(ws.cell(row=cell.row, column=end_col).value == None):
                                end_col = end_col + 1

                            is_seasonality_given = False
                            sub_span_list = list()
                            position = 0

                            for i in range(start_col, end_col):

                                if ws.cell(row=cell.row + 1, column=i).value != None:
                                    if len(sub_span_list) > 0:
                                        sub_spans.append({
                                            "namelist": sub_span_list,
                                            "position": position
                                            })
                                    is_seasonality_given = False
                                    sub_span_list = list()
                                    position += 1
                                if cell.row + 2 < corp_framework_coords_type2["start_row"] and ws.cell(row=cell.row + 2, column=i).value != None :
                                    fare_class_text = ws.cell(row=cell.row + 1, column=i).value.strip() if ws.cell(row=cell.row + 1, column=i).value != None else fare_class_text
                                    fare_class_text = re.sub('\s+(-\s+)?', ' ', fare_class_text)
                                    is_seasonality_given = True

                                if is_seasonality_given:
                                    seasonality_text = ws.cell(row=cell.row + 2, column=i).value.strip()
                                    seasonality_text = re.sub('\s+(-\s+)?', ' ', seasonality_text)
                                    fare_class = fare_class_text + "||" + seasonality_text
                                    if fare_class_text not in fare_brand_list:
                                        fare_brand_list.append(fare_class_text)
                                    sub_span_list.append(seasonality_text)
                                else:
                                    fare_class = ws.cell(row=cell.row + 1, column=i).value.strip()
                                    fare_class = re.sub('\s+(-\s+)?', ' ', fare_class)
                                    if fare_class not in fare_brand_list:
                                        fare_brand_list.append(fare_class)
                                # fare_class = re.sub('\s+(-\s+)?', ' ', fare_class)
                                discount_percent_cols[fare_class] = {'col': get_column_letter(i)}
                                if 'start_row' not in corp_framework_coords_type2:
                                    corp_framework_coords_type2['start_row'] = cell.row + 1

                            if len(sub_span_list) > 0:
                                sub_spans.append({
                                    "namelist": sub_span_list,
                                    "position": position
                                    })
                            fares_discount_coords_set = True
                
                if origin_destination_coords_type2_set:
                    for comp in components_type2_regex_patterns:
                        for pattern in components_type2_regex_patterns[comp]:
                            if comp not in components_coords_type2 and (re.search(pattern, value, flags=re.IGNORECASE) is not None):
                                components_coords_type2[comp] = {'col': get_column_letter(cell.column)}
                                component_list.append(comp)
        
    corp_framework_coords_type2['fare_family'] = fare_family_cols
    corp_framework_coords_type2['components'] = components_coords_type2
    corp_framework_coords_type2['discount_percent'] = discount_percent_cols

    print("header details")
    header_details = dict()
    baseheaders.extend(component_list)
    header_details["baseheaders"] = baseheaders
    header_details["base_spans"] = list()
    header_details["base_spans"].append(fare_family_list)
    header_details["base_spans"].append(fare_brand_list)
    if len(sub_spans) > 0:
        header_details["sub_spans"] = sub_spans
        header_details["seasonality"] = True
    
    pprint.pprint(header_details)

    return corp_framework_coords_type2, header_details

def parse_corporate_sheet_type2(sheet, coords):
    
    complete_sheet_data = list()
    data_list = list()
    cur_row = coords['start_row']
    prev_origin = None
    prev_destination = None
    prev_component = dict()
    
    while(True):
        value_dict = dict()
        row_str = str(cur_row)
        
        origin_coord = coords['origin']['col'] + str(row_str)
        destination_coord = coords['destination']['col'] + str(row_str)
        
        origin_val = sheet[origin_coord].value
        destination_val = sheet[destination_coord].value
        
        if destination_val == None:
            break_flag = True
            for key in coords['components']:
                coord = coords['components'][key]['col'] + str(row_str)
                if (sheet[coord].value != None):
                    break_flag = False
            if break_flag:
                break
        value_dict['origin'] = origin_val if origin_val != None else prev_origin
        value_dict['destination'] = destination_val if destination_val != None else prev_destination
        value_dict['discount_details'] = list()
        
        if origin_val != None:
            prev_origin = origin_val
        if destination_val != None:
            prev_destination = destination_val
        
        for key in coords['discount_percent']:
            coord = coords['discount_percent'][key]['col'] + str(row_str)
            discount_raw = sheet[coord].value
            if isinstance(discount_raw, float):
                discount_raw = round(discount_raw * 100, 1) 
            value_dict['discount_details'].append({
                'discount_percent_class': key,
                'discount_percent_value': discount_raw
            })
        
        for key in coords['components']:
            coord = coords['components'][key]['col'] + str(row_str)
            if key not in prev_component:
                prev_component[key] = sheet[coord].value
            value_dict[key] = sheet[coord].value if sheet[coord].value != None else prev_component[key]
            
        
        value_dict['fare_family_details'] = list()
        for key in coords['fare_family']:
            coord = coords['fare_family'][key]['col'] + str(row_str)
            value_dict['fare_family_details'].append({
                'fare_family_class': key,
                'fare_family_value': sheet[coord].value
            })
            
        print(value_dict)
        complete_sheet_data.append(value_dict)
        print('*******' * 10)
        cur_row = cur_row + 1

    return complete_sheet_data

def parse_fares(excel_path, corporate_name, raw_db, SYSTEM_DATE, SYSTEM_TIME, FILE_DATE, FILE_TIME, fare_sheets=None):

    wb = load_workbook(excel_path)
    complete_excel_data = dict()
    complete_excel_data['corporate_name']  = corporate_name
    # complete_excel_data['corporate_name']  = "test"
    complete_excel_data['system_date'] = SYSTEM_DATE
    complete_excel_data['system_time'] = SYSTEM_TIME
    complete_excel_data['file_date'] = FILE_DATE
    complete_excel_data['file_time'] = FILE_TIME
    complete_excel_data['sheets'] = list()

    sheet_names = wb.sheetnames
    sheet_names = fare_sheets

    sheet_type = 0
     # discount_applied_fare_family_regex_patterns = [
    # "fare\\s*family\\s*to\\s*which\\s*the\\s*discount\\s*to\\s*be\\s*applied",
    # "discounts\\s*applicable\\s*on\\s*published\\s*business\\s*and\\s*economy\\s*class\\s*fare\\s*basis"
    # ]
    # print(discount_applied_fare_family_regex_patterns)
    discount_applied_fare_family_regex_patterns = raw_db.corporate.patterns.find_one(
            {"flag": "keyword_pattern"}, 
            {"fare_sheet.type2_discount_applied_fare_family_regex_patterns": 1, "_id": 0})["fare_sheet"]["type2_discount_applied_fare_family_regex_patterns"]
    # print(discount_applied_fare_family_regex_patterns)

    # discount_percentage_regex_patterns = ["discount\\s*percentage\\s*(/\\s*)?fixed\\s*amount", "percentage\\s*discount$"]
    # print(discount_percentage_regex_patterns)
    discount_percentage_regex_patterns = raw_db.corporate.patterns.find_one(
            {"flag": "keyword_pattern"},
            {"fare_sheet.type2_discount_percentage_regex_patterns": 1, "_id": 0})["fare_sheet"]["type2_discount_percentage_regex_patterns"]
    # print(discount_percentage_regex_patterns)    

    for sheet_name in sheet_names:
        cur_sheet_data = dict()
        cur_sheet_data['sheet_name'] = sheet_name
        ws = wb[sheet_name]
        is_fare_sheet = True
        if is_fare_sheet:
            print('sheet name :', sheet_name)
            ws = wb[sheet_name]
            for row_cells in ws.iter_rows():
                for cell in row_cells:
                    value = cell.value
                    if isinstance(value, str):
                        for pattern in discount_applied_fare_family_regex_patterns:
                            if (re.search(pattern, value, flags=re.IGNORECASE) is not None):
                                sheet_type = 2
                                break

                        for pattern in discount_percentage_regex_patterns:
                            if (re.search(pattern, value, flags=re.IGNORECASE) is not None):
                                sheet_type = 2
                                break

                    if sheet_type != 0:
                        break

                if sheet_type != 0:
                    break

            if sheet_type == 0:
                sheet_type = 1
            coords_def = None
            print("sheet_type : " + str(sheet_type))
            if sheet_type == 1:
                coords_def = get_coords_def_type1(ws, raw_db)
                cur_sheet_data["category25_type"] = "type1"
            else:
                coords_def, header_detail = get_coords_def_type2(ws, raw_db)
                cur_sheet_data['header_detail'] = header_detail 
                cur_sheet_data["category25_type"] = "type2"
            print('coordinates definition : ')
            print('****' * 10)
            pprint.pprint(coords_def)
            print('****' * 10)

            if 'sheet_desc' in coords_def:
                cur_sheet_data['sheet_desc'] = coords_def['sheet_desc']
            if sheet_type == 1:
                sheet_data, table_data  = parse_corp_framework_sheet_type1(ws, coords_def)
                cur_sheet_data['sheet_data'] = sheet_data
                cur_sheet_data['table_data'] = table_data
            else:
                sheet_data  = parse_corporate_sheet_type2(ws, coords_def)
                cur_sheet_data['sheet_data'] = sheet_data
                pass
            complete_excel_data['sheets'].append(cur_sheet_data)
    
    corporate_framework_col = 'corporate.faresheet'

    insertion_details = raw_db[corporate_framework_col].insert_one(complete_excel_data)
    return insertion_details.inserted_id

if __name__ == '__main__':

    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/AE_Essilor_MEast_2020-21_ZZESL3ZZ_WP239928_New.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Barclays_Global CORP_2019-2022-WP208394.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Ericsson_Global Deal.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/GB_Ineos_NYZ3_LocCorp_WP237164.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/IN_Sandvik_Asia_Mar20-Feb21_NEW_WP240519.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Occidental_2019_-_2021_WP239124.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/SA_CORP_SRACO_KSA_ZZSZZ3ZZ_WP238564_Renewal.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/SC_GLOBAL_SUPPLY_CENTRE_PTY_LTD_NEW_WP240952.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/US_Factset_Research_Systems__Inc._NEW_WP236744.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/DK__AtlasCopCo_Miniglobal_WP237689.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/ES_WP240238__RULES__AB_BIOTICS_WP240238.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/IT_Etro_Spa__Renewal_WP239809.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/MICE_World_Art_Dubai_Agent_WP237107.xlsx'
    
    excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/AU_Western_Australia_Government_-_QGW3_WP240393.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/HK_HSBC_Cugo_NEW_WP239965.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/CN_HCRS_Retail_Design_WP241121_New.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/KR_Hyundai_Construction__NEW_WP240782.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/ZA_ZA20SC006-ExecuJet_Corporate_NEW_WP239471.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Brückner_BKG3_DE CORP-WP228758.xlsx'

    system_date = ""
    system_time = ""
    file_date = ""
    file_time = ""
    fare_sheets = ["Fare records and PCC list"]

    client = get_db_client()
    raw_db = client["rawDB_prod"] 

    parse_fares(excel_path, 'type2test', raw_db, system_date, system_time, file_date, file_time, fare_sheets)