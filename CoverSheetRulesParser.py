import openpyxl as oxl
import re
import datetime
from datetime import datetime
import time
import pymongo
import Utils
# client = pymongo.MongoClient(
#     '3.6.201.161:27022',
#     username='data.EK',
#     password='data@123',
#     authSource='admin',
#     authMechanism='SCRAM-SHA-1'
# )
# db1 = client['rawDB_prod']
logger = Utils.get_logger("Cover sheet rule parser")

def parse_coversheet_rules(db1, excel, corporate_name, system_date, system_time, file_date, file_time, cover_sheet_name=None):
    logger.info('Cover sheet rule parser started')
    st=time.time()
    wb = oxl.load_workbook(excel)
    for sheet in wb:
        index = wb.index(sheet)
        ws = wb.worksheets[index]
        list1 = []
        print(sheet.title)
        # if "Cover Sheet" in sheet.title or "nstructions" in sheet.title:
        if sheet.title == cover_sheet_name:
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value != None:
                        value = cell.value
                        list1.append(value)
            print(list1)
            for i in range(len(list1)):
                if (str(type(list1[i]))) in ["<class 'datetime.datetime'>"]:
                    list1[i] = str(datetime.strftime(list1[i], "%d-%m-%Y"))
                if (str(type(list1[i]))) in ["<class 'long'>"] or (str(type(list1[i]))) in ["<class 'float'>"] or (str(type(list1[i]))) in ["<class 'int'>"]:
                    list1[i] = str(list1[i])
            keys_dict = []
            for value in list1:
                if value.startswith("Comment") or value.startswith("Other Notes") or value.startswith("Renewals") or value.startswith("Notes") or value.startswith("CorporateTerms and Conditions"):
                    keys_dict.append(value)
            output = []
            print(keys_dict)
            if len(keys_dict)>0:
                x=list1.index(keys_dict[0])
                for val in list1[x:]:
                    if val in keys_dict:
                        name=val
                    else:
                        output.append({"field_key":name,"field_value":val})
                        if val.startswith("Travel Origin") :
                            del output[-1]
                            break
                print(output)
                print(len(output))
            else:output=[]

            insertion_details = db1.corporate.rulesheet.insert_one({"corporate_name":corporate_name,"other_notes":output, \
                "system_date": system_date, "system_time": system_time, "file_date": file_date, "file_time": file_time})
            logger.info('Cover sheet rule parser completed')
            return insertion_details.inserted_id
            break
            
    print(time.time()-st)
