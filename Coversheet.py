import time
import pymongo
import openpyxl
from datetime import datetime, timedelta
import re
from time import strptime

def check(a, b):
    if a[0] == b[0]:
        return True
    else:
        return False

def excel1(cover_shee, db1, system_date, system_time, file_date, file_time):
    cover_sheet={"details_of_agreement":{"corporate_name":cover_shee['corporate_name'].replace(u'\xa0', '  '),
                                         "partnership":cover_shee['partnership'],
                                         "travel_period":{"from":cover_shee['travel_period_from'],
                                                          "to":cover_shee['travel_period_to'],
                                 "completed_by":cover_shee["entire_journey_must_be_completed_by"]      
                                                          },
                                         "sales_period":{"from":cover_shee['sales_period_from'],
                                                         "to":cover_shee['sales_period_to']},
                                        "address_of_corporate":cover_shee['address_of_the_corporate'],
                                                         "description":cover_shee['description'],
                                                         "name_and_person_signing_the_corporate":cover_shee["name_and_person_signing_the_corporate"],
                                         "corporate_approval_code":cover_shee['corporate_approval_code'],
                                         "plate_stock_restriction":cover_shee['plate_stock_restriction'],
                                         'nature_of_business':cover_shee['nature_of_business'],
                                         'arc_iata_numbers':cover_shee['arc_iata_numbers'],
                                         'entire_journey_must_be_completed_by':cover_shee['entire_journey_must_be_completed_by']
                                                         },
                 "details_of_ticketing_instructions":{"ticket_designator":cover_shee['ticket_designator'],
                                                   "zonal_approach":cover_shee['zonal_approach'],
                                                   "osi_field_entry_in_pnr":cover_shee['osi_field_entry_in_pnr'],
                                                   "fare_amount_box":cover_shee['fare_amount_box'],
                                                   "ticketing_ttl_osi_field_entry_in_pnr":cover_shee['ticketing_ttl_osi_field_entry_in_pnr'],
                                                   'validating_carrier':cover_shee['validating_carrier'],
                                                   'fare_rules':cover_shee['fare_rules'],
                                                   'ticketed_fare_basis_code':cover_shee['ticketed_fare_basis_code'],
                                                   'child_infant_discounts':cover_shee[ 'child_infant_discounts'],
                                                   'tour_code':cover_shee[ 'tour_code'],
                                                   'disclaimer':cover_shee['disclaimer'],
                                                   "baggage_allowance":cover_shee["baggage_allowance"],
                                                   "gross_nett":cover_shee["gross_nett"],
                                                   "market_comission":cover_shee['market_comission']},
                 "details_required_for_private_filing":{'gds_account_code':cover_shee['gds_account_code'],
                                                        'current_fare_display_pricing_entryused_to_retrieve_corporate_deal':cover_shee['current_fare_display_pricing_entryused_to_retrieve_corporate_deal'],
                                                        'distribution_method':cover_shee['distribution_method'],
                                                        'pcc_gds':cover_shee['pcc_gds'],
                                                        "account_code":cover_shee['account_code'],
                                                        "description_justification_details": cover_shee[
                                                            'description_justification_details'],
                                                            "combinations":cover_shee["combinations"]
                                                        },
                }
    cat_list=[]
    if cover_sheet['details_of_agreement']["sales_period"]['from']==None:
        cover_sheet['details_of_agreement']["sales_period"]['from']="01-01-1900"
    id = db1.corporate.coversheet.insert_one(
            {"corporate_name":cover_sheet["details_of_agreement"]["corporate_name"],"cover_sheet": cover_sheet,
             "system_date": system_date, "system_time": system_time, "file_date": file_date, "file_time": file_time,"cat_list":cat_list})
    return id.inserted_id


def excel_parsing(db1,cover_sheet, excel, system_date, system_time, file_date, file_time, cover_sheet_name=None):
    try:
        cover_shee = {"corporate_name":None,
        "partnership":None,
        "ticket_designator":None,
        "tour_code":None,
        "ticketed_fare_basis_code":None,
        "osi_field_entry_in_pnr":None,
        "zonal_approach":None,
        "fare_amount_box":None,
        "fare_rules":None,
        "validating_carrier":None,
        "gross_nett":None,
        "baggage_allowance":None,
        "child_infant_discounts":None,
        "disclaimer":None,
        "travel_period":None,
        "sales_period":None,
        "sales_period_from":None,
        "sales_period_to":None,
        "travel_period_from":None,
        "travel_period_to":None,
        "address_of_the_corporate":None,
        "name_and_person_signing_the_corporate":None,
        "arc_iata_numbers":None,
        "entire_journey_must_be_completed_by":None,
        "plate_stock_restriction":None,
        "gds_account_code":None,
        "distribution_method":None,
        "pcc_gds":None,
        "account_code":None,
        "description":None,
        "description_justification_details":None,
        "ticketing_ttl_osi_field_entry_in_pnr":None,
        "corporate_approval_code":None,
        "nature_of_business":None,
        "current_fare_display_pricing_entryused_to_retrieve_corporate_deal":None,
           "combinations":None,
           "market_comission":None}
        # excel = "aeeeeeeeeeeeee.xlsx"
        #print("*************************")
        print("loading excel", excel)
        wb = openpyxl.load_workbook(excel)
        #print("completed loading")
        for sheet in wb:
            index = wb.index(sheet)
            ws = wb.worksheets[index]
            list1 = []
            #print(sheet.title)
            # if "Cover Sheet" in sheet.title or "nstructions" in sheet.title or "discount" in sheet.title or "Profile" in sheet.title:
            if sheet.title == cover_sheet_name:
                for row in range(1, ws.max_row+1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value != None:
                            value = [str(row), str(col), cell.value]
                            list1.append(value)
                #print(list1)
                for i in range(len(list1)):
                    if (str(type(list1[i][2]))) in ["<class 'datetime.datetime'>"]:
                        list1[i][2] = datetime.strftime(list1[i][2], "%d-%m-%Y")
                    if (str(type(list1[i][2]))) in ["<class 'long'>"] or (str(type(list1[i][2]))) in ["<class 'float'>"] or (str(type(list1[i][2]))) in ["<class 'int'>"]:
                        list1[i][2] = str(list1[i][2])
                    # print("before encoding", type(list1[i][2]))
                    # list1[i][2] = list1[i][2].encode('ascii', errors='ignore')
                    print(list1[i][2],"gg")
                    if "E52-5" in list1[i][2]:
                        list1[i][2]="03-04-2020"
                    if "H52+5" in list1[i][2]:
                       list1[i][2]="16-04-2020"
                    if "H52+15" in list1[i][2]:
                       list1[i][2]="26-04-2020"
                    if 'CONCATENATE' in list1[i][2]:
                        y=list1[i][2].split('CONCATENATE')[1]
                        y=y.replace("(",'').replace(")","").split(",")
                        print(y)
                        sss=""
                        print(sss,"str")
                        for el in range(len(y)):
                            if y[el] =='"  "' or y[el]=='" "':
                                sss=sss+" "
                            else:
                                print("surya")
                                print(el)
                                print(type(sss))
                                sss=sss+str(ws[y[el]].value)
                        list1[i][2]=sss
                    # elif "+" in list1[i][2]:
                    #     x=list1[i][2].replace("=","").strip().split("+")
                    #     list1[i][2]=ws[x[0]].value
                    #     print(x)
                    #     if (str(type(list1[i][2]))) in ["<class 'datetime.datetime'>"]:
                    #         new_date = list1[i][2] + timedelta(int(x[1]))
                    #         list1[i][2] = datetime.strftime(new_date, "%d-%m-%Y")
                    # elif "-" in list1[i][2] and "=" in list1[i][2]:
                    #     x=list1[i][2].replace("=","").strip().split("-")
                    #     list1[i][2]=ws[x[0]].value
                    #     print(x)
                    #     if (str(type(list1[i][2]))) in ["<class 'datetime.datetime'>"]:
                    #         new_date = list1[i][2] - timedelta(int(x[1]))
                    #         list1[i][2] = datetime.strftime(new_date, "%d-%m-%Y")
                    elif "=" in list1[i][2] and "&" in list1[i][2]:
                        try:
                            print (list1[i][2])
                            list1[i][2] = list1[i][2].replace("=", '')
                            print(list1[i][2])
                            y = list1[i][2].split("&")
                            print(len(y))
                            y[1] = ws[y[1]].value
                            print(y)
                            k = ""
                            for l in range(len(y)):
                                k = k + y[l]
                                print(y[l])
                            list1[i][2] = k.replace("\"", "")
                        except ValueError as e:
                            pass
                    elif "=" in list1[i][2]:
                        try:
                            list1[i][2] = list1[i][2].replace("=", '')
                            list1[i][2] = ws[str(list1[i][2])].value
                            if (str(type(list1[i][2]))) in ["<class 'datetime.datetime'>"]:
                                list1[i][2] = datetime.strftime(list1[i][2], "%d-%m-%Y")
                        except ValueError as e:
                            pass
                    elif "WIE" in list1[i][2]:
                        list1[i][2]=""

                #print(list1)
                for i in range(len(list1)):
                    for keys in cover_sheet:
                        key=keys
                        for values in range(len(cover_sheet[keys])):
                            if (re.search(str(cover_sheet[keys][values]), str(list1[i][2]), flags=re.IGNORECASE) is not None) :
                                if (check(list1[i],list1[i+1])):
                                    if keys=="travel_period" or keys=="sales_period":
                                        if "On or After" in str(list1[i + 1][2]):
                                            y = list1[i + 1][2].split('On or After')
                                            cover_shee[key+"_from"] = y[1].replace("/", "-").strip()
                                        elif "As per base fare rules" in list1[i+1][2]:
                                            cover_shee[key + "_from"]=list1[i+1][2]
                                            cover_shee[key + "_to"]=list1[i+1][2]
                                        else:
                                            if check(list1[i],list1[i+1]):
                                                cover_shee[key+"_from"]=list1[i+1][2].strip()
                                            else:cover_shee[key+"_from"]="14-06-2019"
                                            if check(list1[i+1], list1[i +2]):
                                                if (re.search(str('to'), str(list1[i+2][2]), flags=re.IGNORECASE) is not None) :
                                                    cover_shee[key+"_to"]=list1[i+3][2].strip()
                                                else:cover_shee[key + "_to"] = list1[i + 2][2].strip()
                                    else:
                                        if list1[i + 1][2]!="Tour Code" and list1[i + 1][2]!="Fare Box" and 'Name and Title' not in list1[i+1][2]:
                                            cover_shee[key] = list1[i + 1][2].strip()
                                        if keys=="current_fare_display_pricing_entryused_to_retrieve_corporate_deal" and list1[i+1][2]!='TBA':
                                            list2=list1[i+1:i+7]
                                            print(list2)
                                            dict1={}
                                            for i in range(len(list2)):
                                                for j in range(len(list2)):
                                                    if list2[i][0]==str(int(list2[j][0])-1) and  list2[i][1]==list2[j][1]:
                                                        dict1[list2[i][2]]=list2[j][2]
                                            print(dict1)
                                            cover_shee[key]=dict1
                                        if keys=="ticket_designator":  
                                            if list1[i+1][2]=="Delegates:" and list1[i+3][2]=="Keynote Speakers:":
                                                cover_shee[key]=str(list1[i+2][2].strip()+","+list1[i+4][2].strip())
                                            else:
                                                cover_shee[key]=list1[i+1][2]
                                        if list1[i+1][2]=="(If not Base rule, click here)"or list1[i+1][2]=="If more than 1 location Click here":
                                            if (check(list1[i],list1[i+2])):
                                                cover_shee[key] = list1[i + 2][2].strip()
                                            else:
                                                cover_shee[key] =None
                                else:
                                    if key=="description_justification_details":
                                        #print(list1[i+1][2])
                                        cover_shee[key]=list1[i+1][2]
                                        print(cover_shee[key],key)
                                    if key=="child_infant_discounts":
                                        if ")" in  list1[i][2]:
                                            cover_shee[key]=list1[i][2].split(")")[1]
                                    if keys=="travel_period" or keys=='sales_period':
                                        if '* ON OR AFTER (DD-MMM-YYYY)' in list1[i+1][2] and check(list1[i+1],list1[i+2]):
                                            cover_shee[key + "_from"] = list1[i + 2][2].strip()
                                        else:cover_shee[key + "_from"]=None
                                        if '* ON OR BEFORE (DD-MMM-YYYY)' in list1[i + 3][2] and check(list1[i + 3],
                                                                                                      list1[i + 4]):
                                            cover_shee[key + "_to"] = list1[i + 4][2].strip()
                                        else:
                                            if '* ON OR BEFORE (DD-MMM-YYYY)' in list1[i + 2][2] and check(list1[i + 2],
                                                                                                           list1[
                                                                                                               i + 3]):
                                                cover_shee[key + "_to"] = list1[i + 3][2].strip()
                                            else:
                                                cover_shee[key + "_from"] = None
                break
        print(cover_shee)
        id = excel1(cover_shee, db1, system_date, system_time, file_date, file_time)
        return cover_shee['corporate_name'].replace(u'\xa0', '  '), id
    except Exception as e:
        print('exception occured while reading template excel', e)
        print(cover_shee)
def parse_coversheet(db1, excel, system_date, system_time, file_date, file_time, cover_sheet_name=None):
    st = time.time()
    print("Starded Excel Parsing")
    cover_sheet = {
        "corporate_name":["^corporate\s*name","^corporate","^client\s*name"],
        "partnership":["^\s*partnership"],
        "ticket_designator":["^ticket\s*designator"],
        "tour_code":["^tour\s*code"],
        "ticketed_fare_basis_code":["^ticketed\s*fare\s*basis\s*code"],
        "osi_field_entry_in_pnr":["^osi\s*field\s*entry\s*in\s*pnr"],
        "zonal_approach":["^zonal\s*approach"],
        "fare_amount_box":["^fare\s*amount\s*box",'^face\s*value'],
        "fare_rules":["^fare\s*rules"],
        "validating_carrier":["^validating\s*carrier"],
        "child_infant_discounts":["^child/\s*infant\s*discounts","^child / Infant discount",'^chd\s*discount\s*applicable','Child / Infant discounts are not applicable'],
        "disclaimer":["^disclaimer"],
        "market_comission":['^market\s*commission\s*apply','^market\s*commission'],
        "baggage_allowance":['^baggage\s*allowance'],
        "travel_period":["^travel\s*period","^travel\s*date"],
        "sales_period": ["^ticketing/\s*sales","^sale\s*period"],
        "sales_period_from":["^sales\s*start"],
        "sales_period_to":["^sales\s*end"],
        "travel_period_from":['^travel\s*start'],
        "travel_period_to":['^travel\s*end'],
        "description_justification_details":['justification\s*details'],
        "address_of_the_corporate":["^address\s*of\s*the\s*corporate","^address"],
        "name_and_person_signing_the_corporate":["^name\s*and\s*title\s*of\s*the\s*person\s*signing\s*the\s*contract"],
        "arc_iata_numbers":["^arc/\s*iata\s*numbers"],
        "entire_journey_must_be_completed_by":["entire\s*journey\s*must\s*be\s*complete\s*by","^all\s*travel\s*to\s*be\s*completed\s*on/before"],
        "plate_stock_restriction":["plate/\s*stock\s*restriction"],
        "gds_account_code":["^gds\s*account\s*code"],
        "distribution_method":["^distribution\s*method"],
        "pcc_gds":["^pcc\s*/\s*gds"],
        "account_code":['^account\s*code'],
        "description":['^description'],
        "ticketing_ttl_osi_field_entry_in_pnr":["^ticketing\s*ttl\s*osi\s*field\s*entry\s*in\s*pnr"],
        "corporate_approval_code":["^corporate\s*approval\s*code"],
        "nature_of_business":['^nature\s*of\s*business'],
        "gross_nett":['^gross/\s*Nett'],
        "current_fare_display_pricing_entryused_to_retrieve_corporate_deal":["^current\s*fare\s*display\s*pricing\s*entryused\s*to\s*retrieve\s*corporate\s*deal","fare\s*display\s*pricing\s*entry\s*used\s*to\s*retrieve\s*corporate\s*deal"]
           ,"combinations":["^Combinations"] }
    return excel_parsing(db1,cover_sheet, excel, system_date, system_time, file_date, file_time, cover_sheet_name)
    print("Time taken", time.time()-st)

if __name__ == '__main__':
    st = time.time()
    print("Starded Excel Parsing")
    cover_sheet = {
        "corporate_name":["^corporate\s*name","^corporate","^client\s*name"],
        "partnership":["^\s*partnership"],
        "ticket_designator":["^ticket\s*designator"],
        "tour_code":["^tour\s*code"],
        "ticketed_fare_basis_code":["^ticketed\s*fare\s*basis\s*code"],
        "osi_field_entry_in_pnr":["^osi\s*field\s*entry\s*in\s*pnr"],
        "zonal_approach":["^zonal\s*approach"],
        "fare_amount_box":["^fare\s*amount\s*box"],
        "fare_rules":["^fare\s*rules"],
        "validating_carrier":["^validating\s*carrier"],
        "child_infant_discounts":["^child/\s*infant\s*discount",'^chd\s*discount\s*applicable'],
        "disclaimer":["^disclaimer"],
        "market_comission":['^market\s*commission\s*apply','^market\s*commission'],
        "baggage_allowance":['^baggage\s*allowance'],
        "travel_period":["^travel\s*period","^travel\s*date"],
        "sales_period": ["^ticketing/\s*sales","^sale\s*period"],
        "sales_period_from":["^sales\s*start"],
        "sales_period_to":["^sales\s*end"],
        "travel_period_from":['^travel\s*start'],
        "travel_period_to":['^travel\s*end'],
        "description_justification_details":['justification\s*details'],
        "address_of_the_corporate":["^address\s*of\s*the\s*corporate","^address"],
        "name_and_person_signing_the_corporate":["^name\s*and\s*title\s*of\s*the\s*person\s*signing\s*the\s*contract"],
        "arc_iata_numbers":["^arc/\s*iata\s*numbers"],
        "entire_journey_must_be_completed_by":["entire\s*journey\s*must\s*be\s*complete\s*by"],
        "plate_stock_restriction":["plate/\s*stock\s*restriction"],
        "gds_account_code":["^gds\s*account\s*code"],
        "distribution_method":["^distribution\s*method"],
        "pcc_gds":["^pcc\s*/\s*gds"],
        "account_code":['^account\s*code'],
        "description":['^description'],
    "elgibility":["^eligibility"],
    "applicable_discounts":["^applicable\s*discounts"],
    "other_fares":["^other\s*fares"],
    "upfront_discounts":["^upfront\s*discounts"],
    "commission":["^commission"],
    "rules_conditions":["^rules\s*conditions"],
       "ticketing_ttl_osi_field_entry_in_pnr":["^ticketing\s*ttl\s*osi\s*field\s*entry\s*in\s*pnr"],
        "corporate_approval_code":["^corporate\s*approval\s*code"],
        "nature_of_business":['^nature\s*of\s*business'],
        "gross_nett":['^gross/\s*Nett'],
        "current_fare_display_pricing_entryused_to_retrieve_corporate_deal":["^current\s*fare\s*display\s*pricing\s*entryused\s*to\s*retrieve\s*corporate\s*deal","fare\s*display\s*pricing\s*entry\s*used\s*to\s*retrieve\s*corporate\s*deal"],
        "combinations":["^Combinations"]
            }
    client = pymongo.MongoClient(
        '3.6.201.161:27022',
        username='data.EK',
        password='data@123',
        authSource='admin',
        authMechanism='SCRAM-SHA-1'
    )
    db1 = client['rawDB_prod']
    excel_parsing(db1,cover_sheet)
    print("Time taken", time.time()-st)





