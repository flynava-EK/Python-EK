from openpyxl import load_workbook
from pymongo import MongoClient
import re
import pprint

def get_db_client():

    client = MongoClient(
            '3.6.201.161:27022',
            username='data.EK',
            password='data@123',
            authSource='admin',
            #authMechanism='SCRAM-SHA-256'
        )

    return client


def identify_sheets(excel, raw_db):

    name_detail = dict()
    # keywords_record = raw_db.corporate.keywords.find_one({}, {"sheet_identification.sheet_detail": 1})
    # sheet_details = list()
    # try:
    #     sheet_details = keywords_record["sheet_identification"]["sheet_detail"]
    # except Exception as e:
    #     raise(e)
    sheet_details = [
        {
            "sheet_name" : "fare_sheet",
            "patterns" : ['global\s*corporate\s*template', '^corporate\s*framework\s*$', r'local\s*corporate\s*template\s*', \
                          'corporate\s*discounts', r'global\s*corporate\s*\n', 'corporate\s*framework\s*\n', 'all promotional fares on sale for a limited time period are excluded from any discount', \
                         r'deal\s*offer\s*\n', "origin\s*$", "destination\s*$", "online & interline", 'fare\s*family\s*to\s*which\s*the\s*discount\s*to\s*be\s*applied\s*$'],
            "threshold" : 2
        },
        {
            "sheet_name": "cover_sheet",
            "patterns": ["corporate\s*name", "details\s*of\s*agreement", "client\s*name", "details\s*of\s*ticketing\s*instructions"],
            "threshold": 2
        },
        {
            "sheet_name": "iata_sheet",
            "patterns": ["master\s*distribution\s*list", "iata\s*code", "agent\s*name", "agency\s*name", "dedicated\s*iata", "agency$", \
                         "iata\s*number"],
            "threshold": 2
        },
        {
            "sheet_name": "rule_sheet",
            "patterns": ["fares\s*&\s*ticketing\s*particulars", "markets\s*with\s*rule\s*waiver"],
            "threshold": 1
        }
    ]


    wb = load_workbook(excel)
    sheet_names = wb.sheetnames
    # sheet_names = ["Notes"]
    print(sheet_names)

    for sheet in sheet_names:
        ws = wb[sheet]
        if ws.sheet_state == "hidden": 
            continue
        matched = False

        cur_thresholds = {
            "cover_sheet": 0,
            "fare_sheet": 0,
            "iata_sheet": 0,
            "rule_sheet": 0
        }

        for row in ws.iter_rows():
            for cell in row:
                value = cell.value

                if isinstance(value, str):
                    # print(value)
                    # print("****" * 10)
                    for detail in sheet_details:
                        sheet_name = detail["sheet_name"]
                        patterns = detail["patterns"]
                        threshold = detail["threshold"]
                        for pattern in patterns:
                            if (re.search(pattern, value, flags=re.IGNORECASE) is not None):
                                # print(value, pattern)
                                # print("***" * 10)
                                cur_thresholds[sheet_name] += 1
                        if cur_thresholds[sheet_name] >= threshold:
                            if sheet_name not in name_detail:
                                name_detail[sheet_name] = list()
                            if sheet_name == "fare_sheet" and "Rules Exception" in sheet:
                                continue 
                            if sheet not in name_detail[sheet_name]:
                                name_detail[sheet_name].append(sheet)
                            matched = True
                            # break

            #     if matched:
            #         break
            # if matched:
            #     break

        print(sheet)
        pprint.pprint(cur_thresholds)
    pprint.pprint(name_detail)
    return name_detail

if __name__ == "__main__":

    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/AE_Essilor_MEast_2020-21_ZZESL3ZZ_WP239928_New.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Barclays_Global CORP_2019-2022-WP208394.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Ericsson_Global Deal.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/GB_Ineos_NYZ3_LocCorp_WP237164.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/IN_Sandvik_Asia_Mar20-Feb21_NEW_WP240519.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Occidental_2019_-_2021_WP239124.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/SA_CORP_SRACO_KSA_ZZSZZ3ZZ_WP238564_Renewal.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/SC_GLOBAL_SUPPLY_CENTRE_PTY_LTD_NEW_WP240952.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/US_Factset_Research_Systems__Inc._NEW_WP236744.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/DK__AtlasCopCo_Miniglobal_WP237689.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/ES_WP240238__RULES__AB_BIOTICS_WP240238.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/IT_Etro_Spa__Renewal_WP239809.xlsx'
    excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/MICE_World_Art_Dubai_Agent_WP237107.xlsx'


    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/AU_Western_Australia_Government_-_QGW3_WP240393.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/HK_HSBC_Cugo_NEW_WP239965.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/CN_HCRS_Retail_Design_WP241121_New.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/KR_Hyundai_Construction__NEW_WP240782.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/ZA_ZA20SC006-ExecuJet_Corporate_NEW_WP239471.xlsx'
    # excel_path = 'D:/flynava/EK/cat_25/Cat_25_21 templates/Cat_25_21 templates/Brückner_BKG3_DE CORP-WP228758.xlsx'
    
    client = get_db_client()
    raw_db = client["rawDB_prod"]
    identify_sheets(excel_path, raw_db)